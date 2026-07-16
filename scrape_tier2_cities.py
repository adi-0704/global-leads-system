# -*- coding: utf-8 -*-
import os, sys
os.environ.setdefault("PYTHONIOENCODING", "utf-8")
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
"""
scrape_tier2_cities.py
======================
Targets Tier-2 / Tier-3 Indian cities where clinics are LESS likely
to have a website — resulting in a much richer outreach list.

Cities: 32 smaller cities across India
Specialties: Dermatology + Dental
Output: results/tier2_no_website_india.xlsx (new tabs merged into FINAL)
"""

import asyncio, re, time, csv

sys.path.insert(0, os.path.dirname(__file__))
from scraper import GoogleMapsScraper, BusinessRecord, EmailFinder
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Tier-2 / Tier-3 cities (high no-website probability)
# ─────────────────────────────────────────────────────────────────────────────
TIER2_CITIES = [
    # Uttar Pradesh
    "Agra", "Varanasi", "Meerut", "Kanpur", "Prayagraj", "Gorakhpur",
    # Rajasthan
    "Jodhpur", "Udaipur", "Kota", "Ajmer",
    # Gujarat
    "Rajkot", "Bhavnagar", "Anand",
    # Maharashtra
    "Nashik", "Aurangabad", "Kolhapur", "Solapur",
    # Tamil Nadu
    "Madurai", "Trichy", "Salem", "Tirunelveli",
    # Karnataka
    "Mysore", "Mangalore", "Hubli",
    # Andhra Pradesh
    "Vijayawada", "Tirupati", "Guntur",
    # Odisha / Jharkhand / Chhattisgarh
    "Bhubaneswar", "Ranchi", "Raipur",
    # Punjab / Uttarakhand
    "Amritsar", "Ludhiana", "Dehradun",
    # Others
    "Guwahati", "Siliguri", "Thiruvananthapuram",
]

SPECIALTIES = [
    ("dermatologist",   "derma"),
    ("skin clinic",     "derma"),
    ("dental clinic",   "dental"),
    ("dentist",         "dental"),
]

MAX_PER_SEARCH = 20
SEARCH_TIMEOUT = 180
OUTPUT_DIR  = "results"
OUTPUT_XLSX = os.path.join(OUTPUT_DIR, "tier2_no_website_india.xlsx")
OUTPUT_CSV  = os.path.join(OUTPUT_DIR, "tier2_no_website_india.csv")
FINAL_XLSX  = os.path.join(OUTPUT_DIR, "no_website_clinics_india_FINAL.xlsx")

CITY_COLORS = {
    "Agra":"880E4F","Varanasi":"4A148C","Meerut":"1A237E","Kanpur":"BF360C",
    "Prayagraj":"006064","Gorakhpur":"1B5E20","Jodhpur":"E65100",
    "Udaipur":"4E342E","Kota":"37474F","Ajmer":"F57F17","Rajkot":"0D47A1",
    "Bhavnagar":"6A1B9A","Anand":"2E7D32","Nashik":"C62828","Aurangabad":"AD1457",
    "Kolhapur":"558B2F","Solapur":"00695C","Madurai":"E65100","Trichy":"1565C0",
    "Salem":"283593","Tirunelveli":"4527A0","Mysore":"2E7D32","Mangalore":"880E4F",
    "Hubli":"BF360C","Vijayawada":"0D47A1","Tirupati":"6A1B9A","Guntur":"37474F",
    "Bhubaneswar":"00838F","Ranchi":"C62828","Raipur":"1B5E20",
    "Amritsar":"1565C0","Ludhiana":"AD1457","Dehradun":"558B2F",
    "Guwahati":"F57F17","Siliguri":"4527A0","Thiruvananthapuram":"006064",
}

THIN = Side(style="thin", color="DDDDDD")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="0D1117")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

BAD_EMAIL_DOMAINS = ["sentry","wixpress","wix.com","example.com","schema.org","w3.org","google.com"]

def clean_phone(p):
    if not p or str(p) in ('None',''):
        return ""
    digits = re.sub(r"[^\d]", "", str(p))
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    elif len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    return digits[-10:] if len(digits) >= 10 else digits

def clean_name(n):
    if not n or str(n) in ('None',''):
        return ""
    n = str(n).strip()
    n = re.sub(r'\s*\|\s*$', '', n)
    return re.sub(r'\s{2,}', ' ', n).strip()

def clean_email(e):
    if not e or str(e) in ('None',''):
        return ""
    if any(b in str(e).lower() for b in BAD_EMAIL_DOMAINS):
        return ""
    return str(e).strip().lower()

def is_noise(name):
    return str(name).strip().lower() in ("","results","more results","advertisement")

def smart_dedup(records):
    by_phone, no_phone = {}, []
    for rec in records:
        ph = clean_phone(rec.phone)
        if ph:
            if ph not in by_phone:
                by_phone[ph] = rec
            else:
                ex = by_phone[ph]
                if (bool(rec.email)+bool(rec.address)) > (bool(ex.email)+bool(ex.address)):
                    by_phone[ph] = rec
        else:
            no_phone.append(rec)
    by_na = {}
    for rec in no_phone:
        key = (clean_name(rec.name).lower()[:40], rec.address.lower()[:30])
        if key not in by_na:
            by_na[key] = rec
    result = list(by_phone.values()) + list(by_na.values())
    result.sort(key=lambda r: clean_name(r.name).lower())
    return result

def print_banner(t):
    print("\n" + "="*68)
    print(f"  {t}")
    print("="*68)

def print_progress(done, total, start, label=""):
    pct = int(done/total*100) if total else 0
    elapsed = time.time()-start
    speed = done/elapsed if elapsed > 0 else 0
    bar = "#"*int(30*pct/100) + "."*(30-int(30*pct/100))
    eta = (total-done)/speed if speed > 0 else 0
    print(f"\r  [{bar}] {pct:>3}%  {done:>3}/{total}  {speed:.1f}/s ETA:{eta:.0f}s {label}",
          end="", flush=True)

# ── Single search coroutine ────────────────────────────────────────────────
async def _single_search(scraper, query, city, tag, start):
    batch, total_seen = [], 0
    async for record in scraper.search(query, city, max_results=MAX_PER_SEARCH):
        total_seen += 1
        if not record.website and not is_noise(record.name):
            record.specialty = "Dental" if tag == "dental" else "Dermatology"
            batch.append(record)
        print_progress(total_seen, MAX_PER_SEARCH, start,
                       f"Seen:{total_seen} No-site:{len(batch)}")
    return batch, total_seen

# ── Main scraping ──────────────────────────────────────────────────────────
async def run_scrape():
    SEARCH_PLAN = [(sp, city, tag)
                   for city in TIER2_CITIES
                   for (sp, tag) in SPECIALTIES]
    total_searches = len(SEARCH_PLAN)
    print_banner(f"TIER-2/3 CITIES SCRAPER — {len(TIER2_CITIES)} cities x {len(SPECIALTIES)} searches = {total_searches} total")

    city_records = {c: [] for c in TIER2_CITIES}

    async with GoogleMapsScraper(headless=True, delay_ms=900) as scraper:
        for idx, (query, city, tag) in enumerate(SEARCH_PLAN, 1):
            print_banner(f"Search {idx}/{total_searches}: '{query}' in {city} [{tag}]")
            start = time.time()
            batch, seen = [], 0
            try:
                batch, seen = await asyncio.wait_for(
                    _single_search(scraper, query, city, tag, start),
                    timeout=SEARCH_TIMEOUT
                )
            except asyncio.TimeoutError:
                print(f"\n  [TIMEOUT] {SEARCH_TIMEOUT}s -- moving on")
            except Exception as e:
                print(f"\n  [WARN] {e}")
            print(f"\n  --> {len(batch)} no-website (from {seen} results)")
            city_records[city].extend(batch)

    return city_records

# ── Excel helpers ──────────────────────────────────────────────────────────
def style_sheet(ws, city, records):
    clr = CITY_COLORS.get(city, "1A1A2E")
    headers = ["#","Clinic / Doctor Name","Phone","Email","Address","Specialty","City"]
    widths  = [4, 42, 16, 36, 48, 14, 12]

    ws.insert_rows(1); ws.insert_rows(1)
    t = ws.cell(1,1)
    t.value = f"[{city}]  Dental & Dermatology — NO WEBSITE — Tier-2 City Outreach List"
    t.font  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill  = PatternFill("solid", fgColor=clr)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(headers))
    ws.row_dimensions[1].height = 30

    d = sum(1 for r in records if r.specialty=="Dental")
    s = ws.cell(2,1)
    s.value = (f"  Total: {len(records)}  |  Dental: {d}  |  Dermatology: {len(records)-d}  |  "
               f"Generated: {time.strftime('%d %b %Y %H:%M')}")
    s.font  = Font(name="Calibri", size=10, color="CCCCCC", italic=True)
    s.fill  = PatternFill("solid", fgColor="161B22")
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=len(headers))
    ws.row_dimensions[2].height = 20

    for ci,(h,w) in enumerate(zip(headers,widths),1):
        c = ws.cell(3,ci)
        c.value=h; c.font=HEADER_FONT; c.fill=HEADER_FILL
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.border=THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[3].height=22

    ODD=PatternFill("solid",fgColor="F8F9FA")
    EVEN=PatternFill("solid",fgColor="FFFFFF")
    HILITE=PatternFill("solid",fgColor="E8F5E9")

    for ri, rec in enumerate(records, 4):
        em = clean_email(rec.email)
        fill = HILITE if em else (ODD if ri%2==0 else EVEN)
        vals = [ri-3, clean_name(rec.name), clean_phone(rec.phone),
                em, rec.address.strip(), rec.specialty, city]
        for ci,val in enumerate(vals,1):
            cell=ws.cell(ri,ci)
            cell.value=val; cell.fill=fill; cell.border=THIN_BORDER
            cell.alignment=Alignment(vertical="center",wrap_text=(ci in (2,4)))
            if ci==1:
                cell.font=Font(name="Calibri",color="888888",size=9)
                cell.alignment=Alignment(horizontal="center",vertical="center")
            elif ci==2:
                cell.font=Font(name="Calibri",bold=True,size=10)
            elif ci==3:
                cell.font=Font(name="Calibri",color="0D47A1",size=10)
                cell.alignment=Alignment(horizontal="center",vertical="center")
            elif ci==4:
                cell.font=Font(name="Calibri",color="1B5E20",size=10)
            elif ci==6:
                sc="1565C0" if val=="Dental" else "6A1B9A"
                cell.font=Font(name="Calibri",color=sc,size=10,italic=True)
                cell.alignment=Alignment(horizontal="center",vertical="center")
            else:
                cell.font=Font(name="Calibri",size=10)
        ws.row_dimensions[ri].height=20

    ws.freeze_panes="B4"
    ws.auto_filter.ref=f"A3:{get_column_letter(len(headers))}{3+len(records)}"

def build_summary_tab(wb, city_final):
    ws = wb.create_sheet("SUMMARY_TIER2", 0)
    for c,w in zip("ABCDEF",[24,12,12,12,12,20]):
        ws.column_dimensions[c].width=w

    t=ws.cell(1,1)
    t.value="India Tier-2/3 Cities — Clinics With NO Website — Outreach List"
    t.font=Font(name="Calibri",bold=True,size=14,color="FFFFFF")
    t.fill=PatternFill("solid",fgColor="0D1117")
    t.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.merge_cells("A1:F1"); ws.row_dimensions[1].height=32

    sub=ws.cell(2,1)
    sub.value=f"  Generated: {time.strftime('%d %b %Y %H:%M')}  |  32 Tier-2/3 cities scraped"
    sub.font=Font(name="Calibri",size=10,color="CCCCCC",italic=True)
    sub.fill=PatternFill("solid",fgColor="161B22")
    sub.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.merge_cells("A2:F2"); ws.row_dimensions[2].height=20

    for ci,h in enumerate(["City","Total","With Email","Dental","Dermatology","Opportunity"],1):
        c=ws.cell(3,ci); c.value=h; c.font=HEADER_FONT; c.fill=HEADER_FILL
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=THIN_BORDER
    ws.row_dimensions[3].height=22

    ODD=PatternFill("solid",fgColor="F8F9FA")
    EVEN=PatternFill("solid",fgColor="FFFFFF")
    totals=[0,0,0,0]

    for ri,city in enumerate(TIER2_CITIES,4):
        recs=city_final.get(city,[])
        total=len(recs)
        emails=sum(1 for r in recs if clean_email(r.email))
        dental=sum(1 for r in recs if r.specialty=="Dental")
        derma=total-dental
        totals[0]+=total; totals[1]+=emails; totals[2]+=dental; totals[3]+=derma

        fill=ODD if ri%2==0 else EVEN
        opp="HIGH" if total>=8 else ("MEDIUM" if total>=4 else "LOW")
        opp_clr="1B5E20" if total>=8 else ("E65100" if total>=4 else "888888")
        for ci,val in enumerate([city,total,emails,dental,derma,opp],1):
            cell=ws.cell(ri,ci); cell.value=val; cell.fill=fill; cell.border=THIN_BORDER
            cell.alignment=Alignment(horizontal="center" if ci>1 else "left",
                                     vertical="center",indent=1 if ci==1 else 0)
            if ci==1:
                cell.font=Font(name="Calibri",bold=True,
                               color=CITY_COLORS.get(city,"333333"),size=11)
            elif ci==2:
                cell.font=Font(name="Calibri",bold=True,size=13,
                               color="C62828" if total>=6 else "333333")
            elif ci==6:
                cell.font=Font(name="Calibri",color=opp_clr,size=10,
                               bold=(total>=8),italic=True)
            else:
                cell.font=Font(name="Calibri",size=11)
        ws.row_dimensions[ri].height=22

    tr=len(TIER2_CITIES)+4
    for ci,val in enumerate(["TOTAL — Tier-2 Cities",totals[0],totals[1],totals[2],totals[3],
                              f"{int(totals[1]/totals[0]*100)}% have email" if totals[0] else "-"],1):
        c=ws.cell(tr,ci); c.value=val
        c.font=Font(name="Calibri",bold=True,size=13,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor="00C896"); c.border=THIN_BORDER
        c.alignment=Alignment(horizontal="center" if ci>1 else "left",
                              vertical="center",indent=1 if ci==1 else 0)
    ws.row_dimensions[tr].height=28
    ws.freeze_panes="A4"

async def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    start = time.time()

    # Phase 1: Scrape
    city_records = await run_scrape()

    raw_total = sum(len(v) for v in city_records.values())
    print_banner(f"Phase 1 done: {raw_total} raw no-website records")

    # Phase 2: Dedup per city
    print_banner("Cleaning & Deduplicating per City")
    city_final = {}
    for city in TIER2_CITIES:
        raw = city_records[city]
        unique = smart_dedup(raw)
        city_final[city] = unique
        print(f"  {city:<22}: {len(raw):>3} raw -> {len(unique):>3} unique")

    total_unique = sum(len(v) for v in city_final.values())
    print_banner(f"Clean total: {total_unique} unique no-website clinics (Tier-2 cities)")

    # Phase 3: Build tier-2 Excel
    print_banner("Building Tier-2 Excel Workbook")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_summary_tab(wb, city_final)

    for city in TIER2_CITIES:
        recs = city_final[city]
        ws = wb.create_sheet(title=city[:28])
        ws.append(["#","Clinic / Doctor Name","Phone","Email","Address","Specialty","City"])
        for r in recs:
            ws.append([0, r.name, r.phone, r.email, r.address, r.specialty, city])
        style_sheet(ws, city, recs)
        print(f"  Sheet '{city}' -- {len(recs)} records")

    wb.save(OUTPUT_XLSX)
    print(f"\nTier-2 Excel saved: {OUTPUT_XLSX}  ({round(os.path.getsize(OUTPUT_XLSX)/1024,1)} KB)")

    # Phase 4: Master CSV (all records together)
    print("Saving master CSV...")
    with open(OUTPUT_CSV,"w",newline="",encoding="utf-8-sig") as f:
        writer=csv.writer(f)
        writer.writerow(["#","Name","Phone","Email","Address","Specialty","City","Tier"])
        n=0
        for city in TIER2_CITIES:
            for r in city_final[city]:
                n+=1
                writer.writerow([n,clean_name(r.name),clean_phone(r.phone),
                                  clean_email(r.email),r.address.strip(),
                                  r.specialty,city,"Tier-2"])
    print(f"CSV saved: {OUTPUT_CSV}  ({round(os.path.getsize(OUTPUT_CSV)/1024,1)} KB)")

    # Summary
    elapsed = time.time()-start
    print_banner("COMPLETE!")
    print(f"  Total unique clinics (Tier-2) : {total_unique}")
    print(f"  Cities covered                : {len(TIER2_CITIES)}")
    print(f"  Time taken                    : {elapsed/60:.1f} minutes")
    print(f"\n  {'City':<22} {'Total':>6} {'Dental':>7} {'Derma':>8}")
    print("  "+"-"*48)
    for city in TIER2_CITIES:
        recs=city_final[city]
        d=sum(1 for r in recs if r.specialty=="Dental")
        marker="<<<HIGH" if len(recs)>=8 else ""
        print(f"  {city:<22} {len(recs):>6} {d:>7} {len(recs)-d:>8}  {marker}")
    total_d=sum(sum(1 for r in city_final[c] if r.specialty=="Dental") for c in TIER2_CITIES)
    print("  "+"-"*48)
    print(f"  {'TOTAL':<22} {total_unique:>6} {total_d:>7} {total_unique-total_d:>8}")

if __name__ == "__main__":
    asyncio.run(main())
