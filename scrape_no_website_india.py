# -*- coding: utf-8 -*-
import os, sys
os.environ.setdefault("PYTHONIOENCODING", "utf-8")
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
"""
scrape_no_website_india.py
==========================
Scrapes dermatologist + dental clinics across major Indian cities.
FILTERS: Only keeps records with NO website (our target outreach list).
Exports a clean Excel file with one tab per city.

Cities covered:
  Delhi, Mumbai, Pune, Bangalore, Chennai, Hyderabad,
  Ahmedabad, Kolkata, Jaipur, Surat, Lucknow, Chandigarh,
  Nagpur, Indore, Kochi, Coimbatore, Bhopal, Patna,
  Visakhapatnam, Vadodara

Output: results/no_website_clinics_india.xlsx
"""

import asyncio
import os
import sys
import time
import re

sys.path.insert(0, os.path.dirname(__file__))

from scraper import GoogleMapsScraper, BusinessRecord, EmailFinder
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────

MAX_PER_SEARCH  = 20     # per search query (Google Maps ~20 results per search)
SEARCH_TIMEOUT  = 180    # seconds — hard timeout per search to prevent stall
EMAIL_DELAY     = 0.3    # seconds between website visits
OUTPUT_DIR      = "results"
OUTPUT_XLSX     = os.path.join(OUTPUT_DIR, "no_website_clinics_india.xlsx")

BAD_EMAIL_DOMAINS = [
    "sentry", "wixpress", "wix.com", "example.com",
    "schema.org", "w3.org", "google.com", "facebook.com",
    "youtube.com", "instagram.com", "twitter.com",
]

# ─────────────────────────────────────────────────────────────────────────────
# Search Plan: (specialty, area/city) for every city
# ─────────────────────────────────────────────────────────────────────────────

CITIES = [
    "Delhi",
    "Mumbai",
    "Pune",
    "Bangalore",
    "Chennai",
    "Hyderabad",
    "Ahmedabad",
    "Kolkata",
    "Jaipur",
    "Surat",
    "Lucknow",
    "Chandigarh",
    "Nagpur",
    "Indore",
    "Kochi",
    "Coimbatore",
    "Bhopal",
    "Patna",
    "Visakhapatnam",
    "Vadodara",
]

SPECIALTIES = [
    ("dermatologist",    "derma"),
    ("skin clinic",      "derma"),
    ("skin specialist",  "derma"),
    ("dental clinic",    "dental"),
    ("dentist",          "dental"),
    ("dental surgeon",   "dental"),
]

def build_search_plan():
    """Build (query, city, specialty_tag) tuples for all cities × specialties."""
    plan = []
    for city in CITIES:
        for (specialty, tag) in SPECIALTIES:
            plan.append((specialty, city, tag))
    return plan

SEARCH_PLAN = build_search_plan()
print(f"Total searches planned: {len(SEARCH_PLAN)}")

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def is_noise_name(name: str) -> bool:
    stripped = name.strip().lower()
    return stripped in ("", "results", "more results", "advertisement")


def is_bad_email(email: str) -> bool:
    if not email:
        return False
    return any(b in email.lower() for b in BAD_EMAIL_DOMAINS)


def clean_phone(phone: str) -> str:
    """Normalise phone numbers."""
    if not phone:
        return ""
    digits = re.sub(r"[^\d+]", "", phone)
    return digits[:15]


def dedup_records(records: list) -> list:
    """
    Deduplicate by (name.lower + cleaned_phone).
    Prefer record with email over one without.
    """
    best: dict = {}
    for rec in records:
        phone_key = clean_phone(rec.phone)
        name_key  = rec.name.strip().lower()
        key = (name_key, phone_key)
        if key not in best:
            best[key] = rec
        else:
            existing = best[key]
            score_new = bool(rec.email) + bool(rec.address)
            score_old = bool(existing.email) + bool(existing.address)
            if score_new > score_old:
                best[key] = rec
    return list(best.values())


def print_banner(text: str) -> None:
    print("\n" + "=" * 70)
    print(f"  {text}")
    print("=" * 70)


def print_progress(done: int, total: int, start: float, label: str = "Done") -> None:
    pct     = int((done / total) * 100) if total else 0
    elapsed = time.time() - start
    speed   = done / elapsed if elapsed > 0 else 0
    bar_w   = 30
    filled  = int(bar_w * pct / 100)
    bar     = "#" * filled + "." * (bar_w - filled)
    eta     = ((total - done) / speed) if speed > 0 else 0
    print(f"\r  [{bar}] {pct:>3}%  {done:>4}/{total}  {speed:.1f}/s  ETA:{eta:.0f}s  {label}",
          end="", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
# Excel styling helpers
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1A1A2E")          # dark navy
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEAD_FILL  = PatternFill("solid", fgColor="16213E")
ROW_FILL_ODD  = PatternFill("solid", fgColor="F8F9FA")
ROW_FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")
ACCENT_FILL   = PatternFill("solid", fgColor="E8F5E9")          # light green for email rows

THIN = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CITY_COLORS = {
    "Delhi":           "C62828",
    "Mumbai":          "1565C0",
    "Pune":            "6A1B9A",
    "Bangalore":       "2E7D32",
    "Chennai":         "E65100",
    "Hyderabad":       "00838F",
    "Ahmedabad":       "AD1457",
    "Kolkata":         "558B2F",
    "Jaipur":          "F57F17",
    "Surat":           "4527A0",
    "Lucknow":         "00695C",
    "Chandigarh":      "283593",
    "Nagpur":          "BF360C",
    "Indore":          "37474F",
    "Kochi":           "01579B",
    "Coimbatore":      "880E4F",
    "Bhopal":          "1B5E20",
    "Patna":           "E65100",
    "Visakhapatnam":   "0D47A1",
    "Vadodara":        "4A148C",
}


def style_sheet(ws, city: str, records: list) -> None:
    """Apply professional styling to a worksheet."""
    city_color = CITY_COLORS.get(city, "1A1A2E")

    # Column headers
    headers = ["#", "Name", "Phone", "Email", "Address", "Specialty", "City"]
    col_widths = [5, 38, 18, 34, 45, 16, 14]

    # Title row
    ws.insert_rows(1)
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"[{city}] Dental & Dermatology Clinics WITHOUT Website"
    title_cell.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill  = PatternFill("solid", fgColor=city_color)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 32

    # Summary row
    sum_cell = ws.cell(row=2, column=1)
    sum_cell.value = (
        f"  Total: {len(records)} clinics  |  "
        f"With Email: {sum(1 for r in records if r[3])}  |  "
        f"Derma: {sum(1 for r in records if r[5]=='Dermatology')}  |  "
        f"Dental: {sum(1 for r in records if r[5]=='Dental')}  |  "
        f"Generated: {time.strftime('%d %b %Y %H:%M')}"
    )
    sum_cell.font  = Font(name="Calibri", size=10, color="FFFFFF", italic=True)
    sum_cell.fill  = PatternFill("solid", fgColor="2D2D44")
    sum_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.row_dimensions[2].height = 22

    # Header row
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value     = header
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 22

    # Data rows (starting at row 4 because we shifted 2 rows up)
    for row_num, row_data in enumerate(records, 4):
        fill = ROW_FILL_ODD if row_num % 2 == 0 else ROW_FILL_EVEN
        if row_data[3]:  # has email — highlight
            fill = ACCENT_FILL
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value     = value
            cell.fill      = fill
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 4))
            if col_idx == 1:  # # column
                cell.font = Font(name="Calibri", color="888888", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2:  # Name
                cell.font = Font(name="Calibri", bold=True, size=10)
            elif col_idx == 3:  # Phone
                cell.font = Font(name="Calibri", color="0D47A1", size=10)
            elif col_idx == 4:  # Email
                cell.font = Font(name="Calibri", color="1B5E20", size=10)
            elif col_idx == 6:  # Specialty
                spec_color = "1565C0" if row_data[5] == "Dental" else "6A1B9A"
                cell.font = Font(name="Calibri", color=spec_color, size=10, italic=True)
            else:
                cell.font = Font(name="Calibri", size=10)
        ws.row_dimensions[row_num].height = 18

    # Freeze panes at row 4 (below title + header)
    ws.freeze_panes = "A4"

    # Auto-filter on header row
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3 + len(records)}"


def create_summary_sheet(wb, city_data: dict) -> None:
    """Create a top-level SUMMARY sheet."""
    ws = wb.create_sheet("SUMMARY", 0)

    # Title
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 18

    title = ws.cell(row=1, column=1)
    title.value = "India -- Dental & Dermatology Clinics WITHOUT Website"
    title.font  = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    title.fill  = PatternFill("solid", fgColor="0D1117")
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 36

    sub = ws.cell(row=2, column=1)
    sub.value = f"  Generated: {time.strftime('%d %b %Y %H:%M')}  |  These clinics have no website — your best outreach targets"
    sub.font  = Font(name="Calibri", size=10, color="FFFFFF", italic=True)
    sub.fill  = PatternFill("solid", fgColor="161B22")
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 22

    # Header
    headers = ["City", "Total Clinics", "With Email", "Dental", "Dermatology", "Coverage"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value     = h
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER
    ws.row_dimensions[3].height = 22

    total_all = total_email = total_dental = total_derma = 0

    for row_idx, city in enumerate(CITIES, 4):
        recs = city_data.get(city, [])
        total   = len(recs)
        emails  = sum(1 for r in recs if r[3])
        dental  = sum(1 for r in recs if r[5] == "Dental")
        derma   = sum(1 for r in recs if r[5] == "Dermatology")
        pct_em  = f"{int(emails/total*100)}% have email" if total else "—"

        total_all   += total
        total_email += emails
        total_dental += dental
        total_derma += derma

        fill = ROW_FILL_ODD if row_idx % 2 == 0 else ROW_FILL_EVEN
        city_clr = CITY_COLORS.get(city, "333333")
        row_vals = [city, total, emails, dental, derma, pct_em]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value     = val
            cell.fill      = fill
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left",
                                       vertical="center", indent=1 if col_idx == 1 else 0)
            if col_idx == 1:
                cell.font = Font(name="Calibri", bold=True, color=city_clr, size=11)
            elif col_idx == 2:
                cell.font = Font(name="Calibri", bold=True, size=12,
                                 color="C62828" if total > 50 else "1B5E20")
            else:
                cell.font = Font(name="Calibri", size=10)
        ws.row_dimensions[row_idx].height = 20

    # Totals row
    total_row = len(CITIES) + 4
    totals = ["TOTAL (All India)", total_all, total_email, total_dental, total_derma,
              f"{int(total_email/total_all*100)}% have email" if total_all else "—"]
    for col_idx, val in enumerate(totals, 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.value     = val
        cell.font      = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="00C896")
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left",
                                   vertical="center", indent=1 if col_idx == 1 else 0)
    ws.row_dimensions[total_row].height = 26
    ws.freeze_panes = "A4"


# ─────────────────────────────────────────────────────────────────────────────
# Main scraping routine
# ─────────────────────────────────────────────────────────────────────────────

async def _single_search(scraper, query: str, city: str, tag: str, start: float) -> tuple:
    """Run one search query and return (batch_no_website, total_seen)."""
    batch = []
    total_seen = 0
    async for record in scraper.search(query, city, max_results=MAX_PER_SEARCH):
        total_seen += 1
        if not record.website:
            record.specialty = "Dental" if tag == "dental" else "Dermatology"
            batch.append(record)
        print_progress(total_seen, MAX_PER_SEARCH, start,
                       f"Seen:{total_seen} No-site:{len(batch)}")
    return batch, total_seen


async def run_all_searches() -> dict:
    """
    Returns dict: {city: [BusinessRecord, ...]}
    Only records WITHOUT a website are kept.
    """
    city_records: dict = {city: [] for city in CITIES}
    total_searches = len(SEARCH_PLAN)

    async with GoogleMapsScraper(headless=True, delay_ms=1000) as scraper:
        for idx, (query, city, tag) in enumerate(SEARCH_PLAN, 1):
            print_banner(f"Search {idx}/{total_searches}: '{query}' in {city} [{tag}]")
            start = time.time()
            batch = []
            total_seen = 0

            try:
                batch, total_seen = await asyncio.wait_for(
                    _single_search(scraper, query, city, tag, start),
                    timeout=SEARCH_TIMEOUT
                )
            except asyncio.TimeoutError:
                print(f"\n  [TIMEOUT] {SEARCH_TIMEOUT}s limit -- moving on")
            except Exception as e:
                print(f"\n  [WARN] Search failed: {e}")

            print(f"\n  --> {len(batch)} no-website clinics (from {total_seen} total results)")
            city_records[city].extend(batch)

    return city_records


async def find_emails_for_city(records: list) -> None:
    """
    Most no-website records won't have a site to crawl for emails.
    We attempt email discovery via any partial URLs or contact info scraped.
    This is a best-effort phase.
    """
    needs = [r for r in records if r.website and not r.email]
    if not needs:
        return
    finder = EmailFinder(delay=EMAIL_DELAY, check_contact_pages=True)
    for rec in needs:
        try:
            rec.email = finder.find(rec.website) or ""
            if is_bad_email(rec.email):
                rec.email = ""
        except Exception:
            rec.email = ""


async def main() -> None:
    start_total = time.time()

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print_banner("ALL-INDIA No-Website Clinic Scraper — STARTING")
    print(f"  Cities     : {len(CITIES)}")
    print(f"  Specialties: Dermatology + Dental")
    print(f"  Searches   : {len(SEARCH_PLAN)}")
    print(f"  Filter     : Only clinics WITHOUT a website")
    print(f"  Output     : {OUTPUT_XLSX}")

    # ── Phase 1: Scrape ───────────────────────────────────────────────────────
    print_banner("PHASE 1 / 3 — Scraping Google Maps (no-website filter)")
    city_records = await run_all_searches()

    raw_total = sum(len(v) for v in city_records.values())
    print_banner(f"Phase 1 done: {raw_total} no-website clinics across all cities")

    # ── Phase 2: Email discovery (best-effort) ────────────────────────────────
    print_banner("PHASE 2 / 3 — Email Discovery (best-effort from any contact pages)")
    for city, records in city_records.items():
        if records:
            print(f"  {city}: checking {len(records)} records...")
            await find_emails_for_city(records)

    # ── Phase 3: Clean + deduplicate per city ────────────────────────────────
    print_banner("PHASE 3 / 3 — Cleaning & Deduplicating per City")
    clean_city: dict = {}

    for city, records in city_records.items():
        if not records:
            clean_city[city] = []
            continue

        # Remove noise names
        filtered = [r for r in records if not is_noise_name(r.name)]

        # Clean bad emails
        for r in filtered:
            if is_bad_email(r.email):
                r.email = ""
            r.phone = clean_phone(r.phone)

        # Deduplicate
        unique = dedup_records(filtered)
        unique.sort(key=lambda r: r.name.lower())
        clean_city[city] = unique
        print(f"  {city:18}: {len(records):>4} raw -> {len(unique):>4} unique")

    total_unique = sum(len(v) for v in clean_city.values())
    print_banner(f"Clean total: {total_unique} unique no-website clinics")

    # ── Phase 4: Export to Excel ──────────────────────────────────────────────
    print_banner("PHASE 4 / 3 — Building Excel Workbook")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Prepare row data for each city
    city_row_data: dict = {}
    for city, records in clean_city.items():
        rows = []
        for i, rec in enumerate(records, 1):
            specialty = getattr(rec, "specialty", "Dermatology")
            rows.append([
                i,
                rec.name,
                rec.phone,
                rec.email,
                rec.address,
                specialty,
                city,
            ])
        city_row_data[city] = rows

    # Create summary sheet first
    create_summary_sheet(wb, city_row_data)

    # Create one sheet per city
    for city in CITIES:
        records = city_row_data.get(city, [])
        # Tab name: city abbreviation (max 31 chars for Excel)
        tab_name = f"{city}"[:28]
        ws = wb.create_sheet(title=tab_name)

        # Write data rows
        headers = ["#", "Name", "Phone", "Email", "Address", "Specialty", "City"]
        ws.append(headers)  # placeholder — styled below
        for row in records:
            ws.append(row)

        # Apply premium styling
        style_sheet(ws, city, records)

        print(f"  OK Sheet '{tab_name}' -- {len(records)} records")

    # Save workbook
    wb.save(OUTPUT_XLSX)
    file_size_kb = os.path.getsize(OUTPUT_XLSX) / 1024

    # ── Final summary ─────────────────────────────────────────────────────────
    total_time  = time.time() - start_total
    total_email = sum(1 for recs in clean_city.values() for r in recs if r.email)

    print_banner("COMPLETE!")
    print(f"  Total unique clinics (no website) : {total_unique}")
    print(f"  With email discovered              : {total_email}")
    print(f"  Cities covered                     : {len(CITIES)}")
    print(f"  Excel file size                    : {file_size_kb:.1f} KB")
    print(f"  Total time                         : {total_time/60:.1f} minutes")
    print(f"  Output file: {os.path.abspath(OUTPUT_XLSX)}")

    print("\n  PER-CITY BREAKDOWN:")
    print(f"  {'City':<20} {'Total':>6} {'Email':>6} {'Dental':>7} {'Derma':>7}")
    print("  " + "-" * 52)
    for city in CITIES:
        recs = clean_city.get(city, [])
        total   = len(recs)
        emails  = sum(1 for r in recs if r.email)
        dental  = sum(1 for r in recs if r.specialty == 'Dental')
        derma   = total - dental
        print(f"  {city:<20} {total:>6} {emails:>6} {dental:>7} {derma:>7}")
    print("=" * 70)


if __name__ == "__main__":
    asyncio.run(main())
