# -*- coding: utf-8 -*-
"""
postprocess_excel.py
====================
Reads the existing Excel, re-deduplicates with smarter logic
(name + phone + address combo), cleans data, rebuilds Excel
and also saves a master CSV backup.
"""
import os, sys, re, time, csv
os.environ.setdefault("PYTHONIOENCODING", "utf-8")
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_XLSX  = "results/no_website_clinics_india.xlsx"
OUTPUT_XLSX = "results/no_website_clinics_india_FINAL.xlsx"
OUTPUT_CSV  = "results/no_website_clinics_india_FINAL.csv"

CITIES = [
    "Delhi","Mumbai","Pune","Bangalore","Chennai","Hyderabad",
    "Ahmedabad","Kolkata","Jaipur","Surat","Lucknow","Chandigarh",
    "Nagpur","Indore","Kochi","Coimbatore","Bhopal","Patna",
    "Visakhapatnam","Vadodara",
]

CITY_COLORS = {
    "Delhi":"C62828","Mumbai":"1565C0","Pune":"6A1B9A",
    "Bangalore":"2E7D32","Chennai":"E65100","Hyderabad":"00838F",
    "Ahmedabad":"AD1457","Kolkata":"558B2F","Jaipur":"F57F17",
    "Surat":"4527A0","Lucknow":"00695C","Chandigarh":"283593",
    "Nagpur":"BF360C","Indore":"37474F","Kochi":"01579B",
    "Coimbatore":"880E4F","Bhopal":"1B5E20","Patna":"E65100",
    "Visakhapatnam":"0D47A1","Vadodara":"4A148C",
}

def clean_phone(p):
    if not p or str(p) in ('None', ''):
        return ""
    digits = re.sub(r"[^\d]", "", str(p))
    # Normalize Indian numbers: strip leading 0 or 91
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    elif len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    return digits[-10:] if len(digits) >= 10 else digits

def clean_name(n):
    if not n or str(n) in ('None', ''):
        return ""
    # Remove trailing pipe and junk suffixes
    n = str(n).strip()
    n = re.sub(r'\s*\|\s*$', '', n)
    n = re.sub(r'\s{2,}', ' ', n)
    return n.strip()

def clean_email(e):
    if not e or str(e) in ('None', ''):
        return ""
    return str(e).strip().lower()

def clean_address(a):
    if not a or str(a) in ('None', ''):
        return ""
    return str(a).strip()

def smart_dedup(rows):
    """
    Smarter dedup:
    - If phone exists: deduplicate on phone alone (same phone = same clinic)
    - If phone empty: deduplicate on (cleaned_name, first 30 chars of address)
    - Keep the record with the most data filled
    """
    by_phone = {}
    no_phone = []

    for row in rows:
        phone = clean_phone(row[2])
        if phone:
            if phone not in by_phone:
                by_phone[phone] = row
            else:
                # Keep the one with more fields filled
                existing = by_phone[phone]
                score_new = sum(1 for x in row[1:] if x and str(x) not in ('None',''))
                score_old = sum(1 for x in existing[1:] if x and str(x) not in ('None',''))
                if score_new > score_old:
                    by_phone[phone] = row
        else:
            no_phone.append(row)

    # Dedup no-phone records by name+address
    by_name_addr = {}
    for row in no_phone:
        name_key = clean_name(row[1]).lower()[:40]
        addr_key = clean_address(row[4]).lower()[:30]
        key = (name_key, addr_key)
        if key not in by_name_addr:
            by_name_addr[key] = row

    result = list(by_phone.values()) + list(by_name_addr.values())
    # Sort by name
    result.sort(key=lambda r: clean_name(r[1]).lower())
    return result

def read_all_data():
    """Read all data rows from existing Excel."""
    wb = openpyxl.load_workbook(INPUT_XLSX)
    city_data = {}
    for city in CITIES:
        if city not in wb.sheetnames:
            city_data[city] = []
            continue
        ws = wb[city]
        rows = []
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            if row[1] and str(row[1]).strip() not in ('', 'None'):
                rows.append(list(row))
        city_data[city] = rows
    return city_data

THIN = Side(style="thin", color="DDDDDD")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="0D1117")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

def style_sheet(ws, city, records):
    city_color = CITY_COLORS.get(city, "1A1A2E")
    headers = ["#", "Clinic / Doctor Name", "Phone", "Email", "Address", "Specialty", "City"]
    col_widths = [4, 42, 16, 36, 48, 14, 12]

    # Row 1: Title
    ws.insert_rows(1)
    ws.insert_rows(1)
    t = ws.cell(1, 1)
    t.value = f"[{city}]  Dental & Dermatology Clinics — NO WEBSITE  |  Outreach List"
    t.font  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill  = PatternFill("solid", fgColor=city_color)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 30

    # Row 2: Summary
    s = ws.cell(2, 1)
    with_email = sum(1 for r in records if clean_email(r[3]))
    dental  = sum(1 for r in records if str(r[5]) == 'Dental')
    derma   = sum(1 for r in records if str(r[5]) == 'Dermatology')
    s.value = (f"  Total: {len(records)}  |  With Email: {with_email}  |  "
               f"Dental: {dental}  |  Dermatology: {derma}  |  "
               f"Generated: {time.strftime('%d %b %Y %H:%M')}")
    s.font  = Font(name="Calibri", size=10, color="CCCCCC", italic=True)
    s.fill  = PatternFill("solid", fgColor="161B22")
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.row_dimensions[2].height = 20

    # Row 3: Headers
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(3, ci)
        cell.value = h
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 22

    # Data rows from row 4
    ODD  = PatternFill("solid", fgColor="F8F9FA")
    EVEN = PatternFill("solid", fgColor="FFFFFF")
    HILITE = PatternFill("solid", fgColor="E8F5E9")  # has email

    for ri, row in enumerate(records, 4):
        has_email = bool(clean_email(row[3]))
        fill = HILITE if has_email else (ODD if ri % 2 == 0 else EVEN)
        vals = [
            ri - 3,
            clean_name(row[1]),
            clean_phone(row[2]),
            clean_email(row[3]),
            clean_address(row[4]),
            str(row[5]) if row[5] else "",
            str(row[6]) if row[6] else "",
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(ri, ci)
            cell.value  = val
            cell.fill   = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(ci in (2,4)))
            if ci == 1:
                cell.font = Font(name="Calibri", color="888888", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 2:
                cell.font = Font(name="Calibri", bold=True, size=10)
            elif ci == 3:
                cell.font = Font(name="Calibri", color="0D47A1", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 4:
                cell.font = Font(name="Calibri", color="1B5E20", size=10)
            elif ci == 6:
                clr = "1565C0" if val == "Dental" else "6A1B9A"
                cell.font = Font(name="Calibri", color=clr, size=10, italic=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = Font(name="Calibri", size=10)
        ws.row_dimensions[ri].height = 20

    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3+len(records)}"

def build_summary(wb, city_final):
    ws = wb.create_sheet("SUMMARY", 0)
    ws.column_dimensions["A"].width = 24
    for c, w in zip("BCDEF", [12,12,12,12,20]):
        ws.column_dimensions[c].width = w

    # Title
    t = ws.cell(1,1)
    t.value = "India — Dental & Dermatology Clinics With NO Website — Outreach List"
    t.font  = Font(name="Calibri", bold=True, size=15, color="FFFFFF")
    t.fill  = PatternFill("solid", fgColor="0D1117")
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A1:F1"); ws.row_dimensions[1].height = 34

    sub = ws.cell(2,1)
    sub.value = (f"  Generated: {time.strftime('%d %b %Y %H:%M')}  |  "
                 "These clinics have NO website — prime outreach targets for DentalReach")
    sub.font  = Font(name="Calibri", size=10, color="CCCCCC", italic=True)
    sub.fill  = PatternFill("solid", fgColor="161B22")
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A2:F2"); ws.row_dimensions[2].height = 20

    # Header
    for ci, h in enumerate(["City","Total Clinics","With Email","Dental","Dermatology","Notes"],1):
        cell = ws.cell(3, ci)
        cell.value = h; cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 22

    totals = [0,0,0,0]
    ODD  = PatternFill("solid", fgColor="F8F9FA")
    EVEN = PatternFill("solid", fgColor="FFFFFF")

    for ri, city in enumerate(CITIES, 4):
        recs   = city_final.get(city, [])
        total  = len(recs)
        emails = sum(1 for r in recs if clean_email(r[3]))
        dental = sum(1 for r in recs if str(r[5]) == 'Dental')
        derma  = sum(1 for r in recs if str(r[5]) == 'Dermatology')
        totals[0] += total; totals[1] += emails
        totals[2] += dental; totals[3] += derma

        fill = ODD if ri % 2 == 0 else EVEN
        note = "High opportunity" if total >= 6 else ("Medium" if total >= 3 else "Low")
        note_color = "1B5E20" if total >= 6 else ("E65100" if total >= 3 else "888888")
        vals = [city, total, emails, dental, derma, note]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(ri, ci)
            cell.value = val; cell.fill = fill; cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                       vertical="center", indent=1 if ci==1 else 0)
            if ci == 1:
                clr = CITY_COLORS.get(city, "333333")
                cell.font = Font(name="Calibri", bold=True, color=clr, size=11)
            elif ci == 2:
                cell.font = Font(name="Calibri", bold=True, size=13,
                                 color="C62828" if total >= 5 else "333333")
            elif ci == 6:
                cell.font = Font(name="Calibri", color=note_color, size=10, italic=True)
            else:
                cell.font = Font(name="Calibri", size=11)
        ws.row_dimensions[ri].height = 22

    # Total row
    tr = len(CITIES) + 4
    pct = f"{int(totals[1]/totals[0]*100)}% have email" if totals[0] else "-"
    for ci, val in enumerate(["TOTAL — All India", totals[0], totals[1], totals[2], totals[3], pct], 1):
        cell = ws.cell(tr, ci)
        cell.value = val
        cell.font  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        cell.fill  = PatternFill("solid", fgColor="00C896")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                   vertical="center", indent=1 if ci==1 else 0)
    ws.row_dimensions[tr].height = 28
    ws.freeze_panes = "A4"

def main():
    print("Reading existing Excel data...")
    city_data = read_all_data()

    print("Re-deduplicating with smart logic...")
    city_final = {}
    for city in CITIES:
        raw = city_data.get(city, [])
        unique = smart_dedup(raw)
        city_final[city] = unique
        print(f"  {city:<20}: {len(raw):>3} records -> {len(unique):>3} after smart dedup")

    total = sum(len(v) for v in city_final.values())
    print(f"\nTotal after smart dedup: {total}")

    print("\nBuilding final Excel workbook...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_summary(wb, city_final)

    for city in CITIES:
        recs = city_final[city]
        ws   = wb.create_sheet(title=city[:28])
        headers = ["#","Clinic / Doctor Name","Phone","Email","Address","Specialty","City"]
        ws.append(headers)
        for row in recs:
            ws.append(row)
        style_sheet(ws, city, recs)
        print(f"  Sheet '{city}' -- {len(recs)} records")

    wb.save(OUTPUT_XLSX)
    print(f"\nExcel saved: {OUTPUT_XLSX}  ({round(os.path.getsize(OUTPUT_XLSX)/1024,1)} KB)")

    # CSV backup — all cities in one master file
    print("Saving master CSV backup...")
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["#","Name","Phone","Email","Address","Specialty","City"])
        n = 0
        for city in CITIES:
            for row in city_final[city]:
                n += 1
                writer.writerow([n, clean_name(row[1]), clean_phone(row[2]),
                                  clean_email(row[3]), clean_address(row[4]),
                                  str(row[5]) if row[5] else "",
                                  city])
    print(f"CSV saved:   {OUTPUT_CSV}  ({round(os.path.getsize(OUTPUT_CSV)/1024,1)} KB)")

    print("\n" + "="*60)
    print(f"FINAL RESULT: {total} unique no-website clinics across India")
    print("="*60)
    print(f"\n{'City':<22} {'Total':>6} {'Dental':>7} {'Derma':>8} {'Email':>7}")
    print("-"*55)
    for city in CITIES:
        recs = city_final[city]
        d = sum(1 for r in recs if str(r[5])=='Dental')
        e = sum(1 for r in recs if clean_email(r[3]))
        print(f"  {city:<20} {len(recs):>6} {d:>7} {len(recs)-d:>8} {e:>7}")
    print("-"*55)
    total_d = sum(sum(1 for r in city_final[c] if str(r[5])=='Dental') for c in CITIES)
    total_e = sum(sum(1 for r in city_final[c] if clean_email(r[3])) for c in CITIES)
    print(f"  {'TOTAL':<20} {total:>6} {total_d:>7} {total-total_d:>8} {total_e:>7}")

if __name__ == "__main__":
    main()
