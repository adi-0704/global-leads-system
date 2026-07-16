# -*- coding: utf-8 -*-
import os, sys, re, time, csv
os.environ.setdefault("PYTHONIOENCODING", "utf-8")
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

METRO_XLSX  = "results/no_website_clinics_india_FINAL.xlsx"
TIER2_XLSX  = "results/tier2_no_website_india.xlsx"
MASTER_XLSX = "results/MASTER_no_website_india.xlsx"
MASTER_CSV  = "results/MASTER_no_website_india.csv"

METRO_CITIES = [
    "Delhi","Mumbai","Pune","Bangalore","Chennai","Hyderabad",
    "Ahmedabad","Kolkata","Jaipur","Surat","Lucknow","Chandigarh",
    "Nagpur","Indore","Kochi","Coimbatore","Bhopal","Patna",
    "Visakhapatnam","Vadodara",
]
TIER2_CITIES = [
    "Agra","Varanasi","Meerut","Kanpur","Prayagraj","Gorakhpur",
    "Jodhpur","Udaipur","Kota","Ajmer","Rajkot","Bhavnagar","Anand",
    "Nashik","Aurangabad","Kolhapur","Solapur",
    "Madurai","Trichy","Salem","Tirunelveli",
    "Mysore","Mangalore","Hubli",
    "Vijayawada","Tirupati","Guntur",
    "Bhubaneswar","Ranchi","Raipur",
    "Amritsar","Ludhiana","Dehradun",
    "Guwahati","Siliguri","Thiruvananthapuram",
]

ALL_CITIES = METRO_CITIES + TIER2_CITIES

CITY_COLORS = {
    # Metro
    "Delhi":"C62828","Mumbai":"1565C0","Pune":"6A1B9A","Bangalore":"2E7D32",
    "Chennai":"E65100","Hyderabad":"00838F","Ahmedabad":"AD1457","Kolkata":"558B2F",
    "Jaipur":"F57F17","Surat":"4527A0","Lucknow":"00695C","Chandigarh":"283593",
    "Nagpur":"BF360C","Indore":"37474F","Kochi":"01579B","Coimbatore":"880E4F",
    "Bhopal":"1B5E20","Patna":"E65100","Visakhapatnam":"0D47A1","Vadodara":"4A148C",
    # Tier-2
    "Agra":"880E4F","Varanasi":"4A148C","Meerut":"1A237E","Kanpur":"BF360C",
    "Prayagraj":"006064","Gorakhpur":"1B5E20","Jodhpur":"E65100","Udaipur":"4E342E",
    "Kota":"37474F","Ajmer":"F57F17","Rajkot":"0D47A1","Bhavnagar":"6A1B9A",
    "Anand":"2E7D32","Nashik":"C62828","Aurangabad":"AD1457","Kolhapur":"558B2F",
    "Solapur":"00695C","Madurai":"E65100","Trichy":"1565C0","Salem":"283593",
    "Tirunelveli":"4527A0","Mysore":"2E7D32","Mangalore":"880E4F","Hubli":"BF360C",
    "Vijayawada":"0D47A1","Tirupati":"6A1B9A","Guntur":"37474F","Bhubaneswar":"00838F",
    "Ranchi":"C62828","Raipur":"1B5E20","Amritsar":"1565C0","Ludhiana":"AD1457",
    "Dehradun":"558B2F","Guwahati":"F57F17","Siliguri":"4527A0","Thiruvananthapuram":"006064",
}

THIN = Side(style="thin", color="DDDDDD")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="0D1117")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

def cp(p):
    if not p or str(p) in ('None',''): return ""
    d = re.sub(r"[^\d]","",str(p))
    if len(d)==12 and d.startswith("91"): d=d[2:]
    elif len(d)==11 and d.startswith("0"): d=d[1:]
    return d[-10:] if len(d)>=10 else d

def cn(n):
    if not n or str(n) in ('None',''): return ""
    n=re.sub(r'\s*\|\s*$','',str(n).strip())
    return re.sub(r'\s{2,}',' ',n).strip()

def ce(e):
    if not e or str(e) in ('None',''): return ""
    return str(e).strip().lower()

def read_wb_city(wb, city):
    if city not in wb.sheetnames: return []
    ws = wb[city]
    rows=[]
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if row[1] and str(row[1]).strip() not in ('','None'):
            rows.append({
                'name': cn(row[1]),
                'phone': cp(row[2]),
                'email': ce(row[3]),
                'address': str(row[4] or '').strip(),
                'specialty': str(row[5] or ''),
                'city': city,
            })
    return rows

def style_city_sheet(ws, city, records, tier):
    clr = CITY_COLORS.get(city, "1A1A2E")
    headers = ["#","Clinic / Doctor Name","Phone","Email","Address","Specialty","City","Tier"]
    widths  = [4, 40, 15, 35, 46, 13, 14, 8]

    ws.insert_rows(1); ws.insert_rows(1)
    t = ws.cell(1,1)
    d = sum(1 for r in records if r['specialty']=='Dental')
    t.value = f"[{city}]  Dental & Derma — NO WEBSITE — {tier} Outreach List  |  {len(records)} clinics  |  Dental:{d}  Derma:{len(records)-d}"
    t.font  = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    t.fill  = PatternFill("solid", fgColor=clr)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(headers))
    ws.row_dimensions[1].height = 28

    s = ws.cell(2,1)
    s.value = f"  Generated: {time.strftime('%d %b %Y %H:%M')}  |  All records confirmed: NO WEBSITE on Google Maps"
    s.font  = Font(name="Calibri", size=10, color="BBBBBB", italic=True)
    s.fill  = PatternFill("solid", fgColor="161B22")
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=len(headers))
    ws.row_dimensions[2].height = 18

    for ci,(h,w) in enumerate(zip(headers,widths),1):
        c=ws.cell(3,ci); c.value=h; c.font=HEADER_FONT; c.fill=HEADER_FILL
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[3].height=20

    ODD=PatternFill("solid",fgColor="F8F9FA")
    EVEN=PatternFill("solid",fgColor="FFFFFF")
    HILITE=PatternFill("solid",fgColor="E8F5E9")

    for ri,rec in enumerate(records,4):
        fill = HILITE if rec['email'] else (ODD if ri%2==0 else EVEN)
        vals=[ri-3, rec['name'], rec['phone'], rec['email'],
              rec['address'], rec['specialty'], city, tier]
        for ci,val in enumerate(vals,1):
            cell=ws.cell(ri,ci); cell.value=val; cell.fill=fill; cell.border=THIN_BORDER
            cell.alignment=Alignment(vertical="center",wrap_text=(ci in (2,4)))
            if ci==1:
                cell.font=Font(name="Calibri",color="999999",size=9)
                cell.alignment=Alignment(horizontal="center",vertical="center")
            elif ci==2: cell.font=Font(name="Calibri",bold=True,size=10)
            elif ci==3:
                cell.font=Font(name="Calibri",color="0D47A1",size=10)
                cell.alignment=Alignment(horizontal="center",vertical="center")
            elif ci==4: cell.font=Font(name="Calibri",color="1B5E20",size=10)
            elif ci==6:
                sc="1565C0" if val=="Dental" else "6A1B9A"
                cell.font=Font(name="Calibri",color=sc,size=10,italic=True)
                cell.alignment=Alignment(horizontal="center",vertical="center")
            elif ci==8:
                tc="C62828" if val=="Metro" else "1B5E20"
                cell.font=Font(name="Calibri",color=tc,size=9,bold=True)
                cell.alignment=Alignment(horizontal="center",vertical="center")
            else: cell.font=Font(name="Calibri",size=10)
        ws.row_dimensions[ri].height=18

    ws.freeze_panes="B4"
    ws.auto_filter.ref=f"A3:{get_column_letter(len(headers))}{3+len(records)}"

def build_master_summary(wb, all_data):
    ws = wb.create_sheet("MASTER SUMMARY", 0)
    for c,w in zip("ABCDEFGH",[22,10,10,10,10,10,12,20]):
        ws.column_dimensions[c].width=w

    # Title
    t=ws.cell(1,1)
    t.value="INDIA — All Dental & Dermatology Clinics With NO WEBSITE — Complete Outreach List"
    t.font=Font(name="Calibri",bold=True,size=15,color="FFFFFF")
    t.fill=PatternFill("solid",fgColor="0D1117")
    t.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.merge_cells("A1:H1"); ws.row_dimensions[1].height=36

    s=ws.cell(2,1)
    total_all=sum(len(v) for v in all_data.values())
    s.value=(f"  {total_all} unique clinics | 56 cities | Metro (20) + Tier-2/3 (36) | "
             f"Generated: {time.strftime('%d %b %Y %H:%M')}")
    s.font=Font(name="Calibri",size=11,color="CCCCCC",italic=True)
    s.fill=PatternFill("solid",fgColor="161B22")
    s.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.merge_cells("A2:H2"); ws.row_dimensions[2].height=22

    # Section header
    for ci,h in enumerate(["City","Total","Dental","Derma","Email","Tier","Opportunity","State/Region"],1):
        c=ws.cell(3,ci); c.value=h; c.font=HEADER_FONT; c.fill=HEADER_FILL
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=THIN_BORDER
    ws.row_dimensions[3].height=22

    STATE_MAP = {
        "Delhi":"Delhi","Mumbai":"Maharashtra","Pune":"Maharashtra","Bangalore":"Karnataka",
        "Chennai":"Tamil Nadu","Hyderabad":"Telangana","Ahmedabad":"Gujarat","Kolkata":"West Bengal",
        "Jaipur":"Rajasthan","Surat":"Gujarat","Lucknow":"Uttar Pradesh","Chandigarh":"Punjab/HR",
        "Nagpur":"Maharashtra","Indore":"Madhya Pradesh","Kochi":"Kerala","Coimbatore":"Tamil Nadu",
        "Bhopal":"Madhya Pradesh","Patna":"Bihar","Visakhapatnam":"Andhra Pradesh","Vadodara":"Gujarat",
        "Agra":"Uttar Pradesh","Varanasi":"Uttar Pradesh","Meerut":"Uttar Pradesh","Kanpur":"Uttar Pradesh",
        "Prayagraj":"Uttar Pradesh","Gorakhpur":"Uttar Pradesh","Jodhpur":"Rajasthan","Udaipur":"Rajasthan",
        "Kota":"Rajasthan","Ajmer":"Rajasthan","Rajkot":"Gujarat","Bhavnagar":"Gujarat","Anand":"Gujarat",
        "Nashik":"Maharashtra","Aurangabad":"Maharashtra","Kolhapur":"Maharashtra","Solapur":"Maharashtra",
        "Madurai":"Tamil Nadu","Trichy":"Tamil Nadu","Salem":"Tamil Nadu","Tirunelveli":"Tamil Nadu",
        "Mysore":"Karnataka","Mangalore":"Karnataka","Hubli":"Karnataka",
        "Vijayawada":"Andhra Pradesh","Tirupati":"Andhra Pradesh","Guntur":"Andhra Pradesh",
        "Bhubaneswar":"Odisha","Ranchi":"Jharkhand","Raipur":"Chhattisgarh",
        "Amritsar":"Punjab","Ludhiana":"Punjab","Dehradun":"Uttarakhand",
        "Guwahati":"Assam","Siliguri":"West Bengal","Thiruvananthapuram":"Kerala",
    }

    ODD=PatternFill("solid",fgColor="F8F9FA")
    EVEN=PatternFill("solid",fgColor="FFFFFF")
    metro_total=tier2_total=metro_d=tier2_d=0

    # Metro section label
    ml=ws.cell(4,1)
    ml.value="-- METRO CITIES (Tier-1) --"
    ml.font=Font(name="Calibri",bold=True,size=11,color="FFFFFF")
    ml.fill=PatternFill("solid",fgColor="C62828")
    ml.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.merge_cells("A4:H4"); ws.row_dimensions[4].height=20

    ri=5
    for city in METRO_CITIES:
        recs=all_data.get(city,[])
        total=len(recs); d=sum(1 for r in recs if r['specialty']=='Dental')
        em=sum(1 for r in recs if r['email'])
        metro_total+=total; metro_d+=d
        fill=ODD if ri%2==0 else EVEN
        opp="HIGH" if total>=6 else ("MED" if total>=3 else "LOW")
        opp_c="1B5E20" if total>=6 else ("E65100" if total>=3 else "999999")
        for ci,val in enumerate([city,total,d,total-d,em,"Metro",opp,STATE_MAP.get(city,"")],1):
            cell=ws.cell(ri,ci); cell.value=val; cell.fill=fill; cell.border=THIN_BORDER
            cell.alignment=Alignment(horizontal="center" if ci>1 else "left",
                                     vertical="center",indent=1 if ci==1 else 0)
            if ci==1: cell.font=Font(name="Calibri",bold=True,color=CITY_COLORS.get(city,"333"),size=11)
            elif ci==2: cell.font=Font(name="Calibri",bold=True,size=12,color="333333" if total<6 else "C62828")
            elif ci==7: cell.font=Font(name="Calibri",color=opp_c,size=10,italic=True,bold=(total>=6))
            elif ci==6: cell.font=Font(name="Calibri",color="C62828",size=10,bold=True)
            else: cell.font=Font(name="Calibri",size=11)
        ws.row_dimensions[ri].height=20; ri+=1

    # Tier-2 section label
    tl=ws.cell(ri,1)
    tl.value="-- TIER-2 / TIER-3 CITIES --"
    tl.font=Font(name="Calibri",bold=True,size=11,color="FFFFFF")
    tl.fill=PatternFill("solid",fgColor="1B5E20")
    tl.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.merge_cells(f"A{ri}:H{ri}"); ws.row_dimensions[ri].height=20; ri+=1

    for city in TIER2_CITIES:
        recs=all_data.get(city,[])
        total=len(recs); d=sum(1 for r in recs if r['specialty']=='Dental')
        em=sum(1 for r in recs if r['email'])
        tier2_total+=total; tier2_d+=d
        fill=ODD if ri%2==0 else EVEN
        opp="HIGH" if total>=8 else ("MED" if total>=4 else "LOW")
        opp_c="1B5E20" if total>=8 else ("E65100" if total>=4 else "999999")
        for ci,val in enumerate([city,total,d,total-d,em,"Tier-2",opp,STATE_MAP.get(city,"")],1):
            cell=ws.cell(ri,ci); cell.value=val; cell.fill=fill; cell.border=THIN_BORDER
            cell.alignment=Alignment(horizontal="center" if ci>1 else "left",
                                     vertical="center",indent=1 if ci==1 else 0)
            if ci==1: cell.font=Font(name="Calibri",bold=True,color=CITY_COLORS.get(city,"333"),size=11)
            elif ci==2: cell.font=Font(name="Calibri",bold=True,size=12,color="333333" if total<8 else "1B5E20")
            elif ci==7: cell.font=Font(name="Calibri",color=opp_c,size=10,italic=True,bold=(total>=8))
            elif ci==6: cell.font=Font(name="Calibri",color="1B5E20",size=10,bold=True)
            else: cell.font=Font(name="Calibri",size=11)
        ws.row_dimensions[ri].height=20; ri+=1

    # Grand total
    grand=metro_total+tier2_total
    grand_d=metro_d+tier2_d
    for ci,val in enumerate(["GRAND TOTAL — ALL INDIA",grand,grand_d,grand-grand_d,
                              "","56 Cities","",""],1):
        cell=ws.cell(ri,ci); cell.value=val
        cell.font=Font(name="Calibri",bold=True,size=14,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor="00C896"); cell.border=THIN_BORDER
        cell.alignment=Alignment(horizontal="center" if ci>1 else "left",
                                 vertical="center",indent=1 if ci==1 else 0)
    ws.row_dimensions[ri].height=30
    ws.freeze_panes="A4"

def main():
    print("="*65)
    print("  BUILDING MASTER ALL-INDIA OUTREACH EXCEL")
    print("="*65)

    # Read metro data
    print("\nReading metro cities Excel...")
    wb_metro = openpyxl.load_workbook(METRO_XLSX)
    metro_data = {city: read_wb_city(wb_metro, city) for city in METRO_CITIES}
    metro_total = sum(len(v) for v in metro_data.values())
    print(f"  Metro: {metro_total} records across {len(METRO_CITIES)} cities")

    # Read tier-2 data
    print("Reading tier-2 cities Excel...")
    wb_t2 = openpyxl.load_workbook(TIER2_XLSX)
    tier2_data = {city: read_wb_city(wb_t2, city) for city in TIER2_CITIES}
    tier2_total = sum(len(v) for v in tier2_data.values())
    print(f"  Tier-2: {tier2_total} records across {len(TIER2_CITIES)} cities")

    all_data = {**metro_data, **tier2_data}
    grand_total = metro_total + tier2_total
    print(f"\n  GRAND TOTAL: {grand_total} unique no-website clinics across 56 cities")

    # Build master workbook
    print("\nBuilding master Excel...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_master_summary(wb, all_data)

    # Metro sheets
    for city in METRO_CITIES:
        recs = metro_data[city]
        ws = wb.create_sheet(title=f"M-{city}"[:28])
        ws.append(["#","Name","Phone","Email","Address","Specialty","City","Tier"])
        for r in recs: ws.append([0,r['name'],r['phone'],r['email'],r['address'],r['specialty'],city,"Metro"])
        style_city_sheet(ws, city, recs, "Metro")
        print(f"  [Metro]  {city:<20}: {len(recs)} records")

    # Tier-2 sheets
    for city in TIER2_CITIES:
        recs = tier2_data[city]
        ws = wb.create_sheet(title=f"T2-{city}"[:28])
        ws.append(["#","Name","Phone","Email","Address","Specialty","City","Tier"])
        for r in recs: ws.append([0,r['name'],r['phone'],r['email'],r['address'],r['specialty'],city,"Tier-2"])
        style_city_sheet(ws, city, recs, "Tier-2")
        print(f"  [Tier-2] {city:<20}: {len(recs)} records")

    wb.save(MASTER_XLSX)
    sz = round(os.path.getsize(MASTER_XLSX)/1024, 1)
    print(f"\nMaster Excel saved: {MASTER_XLSX}  ({sz} KB)")

    # Master CSV
    print("Saving master CSV...")
    with open(MASTER_CSV,"w",newline="",encoding="utf-8-sig") as f:
        w=csv.writer(f)
        w.writerow(["#","Name","Phone","Email","Address","Specialty","City","Tier"])
        n=0
        for city in ALL_CITIES:
            tier="Metro" if city in METRO_CITIES else "Tier-2"
            for r in all_data[city]:
                n+=1
                w.writerow([n,r['name'],r['phone'],r['email'],r['address'],r['specialty'],city,tier])
    csz=round(os.path.getsize(MASTER_CSV)/1024,1)
    print(f"Master CSV saved:   {MASTER_CSV}  ({csz} KB)")

    print("\n" + "="*65)
    print(f"  DONE! {grand_total} unique no-website clinics")
    print(f"  Metro cities  : {metro_total} records ({len(METRO_CITIES)} cities)")
    print(f"  Tier-2 cities : {tier2_total} records ({len(TIER2_CITIES)} cities)")
    print(f"  Excel size    : {sz} KB")
    print(f"  CSV size      : {csz} KB")
    print("="*65)

if __name__ == "__main__":
    main()
