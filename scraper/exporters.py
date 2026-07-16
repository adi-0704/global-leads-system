"""
exporters.py — CSV, JSON and Excel Export Utilities
====================================================
Converts a list of BusinessRecord objects into downloadable files.

Exported columns (in order):
  Name | Phone | Email | Website | Address | Scraped At

Author  : Google Maps Scraper Project
Requires: openpyxl (for Excel export only)
"""

import csv
import json
import os
from typing import List

from .maps_scraper import BusinessRecord


# ---------------------------------------------------------------------------
# Column definition (single source of truth for all exporters)
# ---------------------------------------------------------------------------

_COLUMNS = ["name", "phone", "email", "website", "address", "scraped_at"]
_HEADERS = ["Name", "Phone", "Email", "Website", "Address", "Scraped At"]


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def export_csv(records: List[BusinessRecord], filepath: str) -> str:
    """
    Write *records* to a CSV file at *filepath*.

    Parameters
    ----------
    records  : List of BusinessRecord objects to export.
    filepath : Destination path (e.g. "results/dentists_mumbai.csv").

    Returns
    -------
    str — Absolute path of the written file.
    """
    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        # utf-8-sig adds BOM so Excel opens without encoding issues
        writer = csv.DictWriter(f, fieldnames=_COLUMNS)
        writer.writerow(dict(zip(_COLUMNS, _HEADERS)))   # human-readable header row
        for rec in records:
            writer.writerow(rec.to_dict())

    return os.path.abspath(filepath)


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------

def export_json(records: List[BusinessRecord], filepath: str) -> str:
    """
    Write *records* to a pretty-printed JSON file at *filepath*.

    Parameters
    ----------
    records  : List of BusinessRecord objects to export.
    filepath : Destination path (e.g. "results/dentists_mumbai.json").

    Returns
    -------
    str — Absolute path of the written file.
    """
    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)

    data = {
        "total": len(records),
        "results": [rec.to_dict() for rec in records],
    }

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return os.path.abspath(filepath)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def export_excel(records: List[BusinessRecord], filepath: str) -> str:
    """
    Write *records* to an Excel (.xlsx) workbook at *filepath*.

    The sheet is auto-formatted:
      - Header row has a dark background with white bold text.
      - Columns are auto-sized to fit content.
      - Alternating row shading for readability.

    Parameters
    ----------
    records  : List of BusinessRecord objects to export.
    filepath : Destination path (e.g. "results/dentists_mumbai.xlsx").

    Returns
    -------
    str — Absolute path of the written file.

    Raises
    ------
    ImportError if openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError("openpyxl is required for Excel export: pip install openpyxl") from exc

    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Results"

    # ── Header row ────────────────────────────────────────────────────────
    header_fill = PatternFill(fill_type="solid", fgColor="1a1a2e")
    alt_fill    = PatternFill(fill_type="solid", fgColor="f0f4ff")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    body_font   = Font(name="Calibri", size=10)
    center      = Alignment(horizontal="center", vertical="center")

    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center

    # ── Data rows ─────────────────────────────────────────────────────────
    for i, rec in enumerate(records, start=2):
        row_data = [
            rec.name, rec.phone, rec.email,
            rec.website, rec.address, rec.scraped_at,
        ]
        ws.append(row_data)
        for cell in ws[i]:
            cell.font = body_font
            cell.alignment = Alignment(vertical="center")
            if i % 2 == 0:
                cell.fill = alt_fill

    # ── Auto-size columns ─────────────────────────────────────────────────
    for col_idx, col_header in enumerate(_HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_header)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    ws.row_dimensions[1].height = 22

    wb.save(filepath)
    return os.path.abspath(filepath)
